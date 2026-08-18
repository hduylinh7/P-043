import io
import logging
import os
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

try:
    from qdrant_client import QdrantClient, models
    try:
        from langchain_qdrant import QdrantVectorStore
    except ImportError:
        from langchain_community.vectorstores import Qdrant as QdrantVectorStore
except ImportError:
    QdrantClient = None
    models = None
    QdrantVectorStore = None

try:
    from langchain_core.documents import Document
    from langchain_core.embeddings import Embeddings
    from langchain_google_genai import GoogleGenerativeAIEmbeddings
    from langchain_text_splitters import RecursiveCharacterTextSplitter
except ImportError:
    Document = None
    Embeddings = object
    GoogleGenerativeAIEmbeddings = None
    RecursiveCharacterTextSplitter = None

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

from src.config import get_settings
from src.db.database import AsyncSessionLocal
from src.repositories.material_repository import MaterialRepository

logger = logging.getLogger(__name__)


class ResilientEmbeddings(Embeddings):
    """Wrapper that tries primary embedding model (Google Generative AI) and falls back to hash-based pseudo-embeddings if primary API fails."""

    def __init__(self, primary: Any):
        self.primary = primary
        self.dim = 3072

    def _hash_vector(self, text: str) -> list[float]:
        import hashlib
        h = hashlib.sha256(text.encode("utf-8")).hexdigest()
        return [(int(h[(i * 2) % len(h) : (i * 2) % len(h) + 2], 16) - 128) / 128.0 for i in range(self.dim)]

    def embed_documents(self, texts: list[str]) -> list[list[float]]:
        try:
            return self.primary.embed_documents(texts)
        except Exception as e:
            logger.warning(f"Primary embedding model failed ({e}). Utilizing resilient hash embeddings fallback.")
            return [self._hash_vector(t) for t in texts]

    def embed_query(self, text: str) -> list[float]:
        try:
            return self.primary.embed_query(text)
        except Exception as e:
            logger.warning(f"Primary query embedding failed ({e}). Utilizing resilient hash query embedding fallback.")
            return self._hash_vector(text)

    def __call__(self, text: str) -> list[float]:
        return self.embed_query(text)


def _get_embedding_function() -> Any:
    """Instantiate Google Generative AI Embeddings with Resilient Embeddings Fallback."""
    settings = get_settings()
    api_key = (
        settings.gemini_api_key
        or settings.google_api_key
        or os.getenv("GEMINI_API_KEY")
        or os.getenv("GOOGLE_API_KEY")
    )
    if not api_key:
        logger.warning("GEMINI_API_KEY / GOOGLE_API_KEY not set. Embedding operations may fail.")

    model_name = settings.embedding_model_name or "models/gemini-embedding-2"
    primary = GoogleGenerativeAIEmbeddings(
        model=model_name,
        google_api_key=api_key,
    )
    return ResilientEmbeddings(primary=primary)


def _get_qdrant_client() -> Any:
    """Get QdrantClient instance connecting to Cloud or local Qdrant."""
    settings = get_settings()
    url = settings.qdrant_url
    api_key = settings.qdrant_api_key or None
    if not url:
        logger.warning("QDRANT_URL is not configured. Falling back to http://localhost:6333")
        url = "http://localhost:6333"
    if QdrantClient is None:
        raise ImportError("qdrant-client package is not installed.")
    return QdrantClient(url=url, api_key=api_key)


def _get_vector_store() -> Any:
    """Get Qdrant vector store instance."""
    settings = get_settings()
    embeddings = _get_embedding_function()

    if QdrantVectorStore is None or QdrantClient is None:
        raise ImportError("Qdrant packages (qdrant-client, langchain-qdrant) are not installed.")

    client = _get_qdrant_client()
    collection_name = settings.qdrant_collection_name or "course_materials"
    try:
        if not client.collection_exists(collection_name):
            client.create_collection(
                collection_name=collection_name,
                vectors_config=models.VectorParams(
                    size=3072,
                    distance=models.Distance.COSINE,
                ),
            )
            logger.info(f"Created Qdrant collection '{collection_name}' successfully.")
    except Exception as e:
        logger.warning(f"Could not verify/create Qdrant collection '{collection_name}': {e}")

    # Ensure Payload Indexes exist for filtering by course_id, material_id, assignment_id, and type on Qdrant
    try:
        client.create_payload_index(
            collection_name=collection_name,
            field_name="metadata.course_id",
            field_schema=models.PayloadSchemaType.KEYWORD,
        )
        client.create_payload_index(
            collection_name=collection_name,
            field_name="metadata.material_id",
            field_schema=models.PayloadSchemaType.KEYWORD,
        )
        client.create_payload_index(
            collection_name=collection_name,
            field_name="metadata.assignment_id",
            field_schema=models.PayloadSchemaType.KEYWORD,
        )
        client.create_payload_index(
            collection_name=collection_name,
            field_name="metadata.type",
            field_schema=models.PayloadSchemaType.KEYWORD,
        )
    except Exception as idx_err:
        logger.debug(f"Payload index check for Qdrant: {idx_err}")

    if hasattr(QdrantVectorStore, "__module__") and QdrantVectorStore.__module__.startswith("langchain_qdrant"):
        return QdrantVectorStore(
            client=client,
            collection_name=collection_name,
            embedding=embeddings,
        )
    else:
        return QdrantVectorStore(
            client=client,
            collection_name=collection_name,
            embeddings=embeddings,
        )


def _ocr_extract_text_from_image(image_bytes: bytes, mime_type: str = "image/png") -> str:
    """Extract text, tables, and math formulas (LaTeX) from image bytes using Gemini Vision API."""
    if not image_bytes or len(image_bytes) < 5120:  # Skip tiny icons/logos < 5KB
        return ""
    try:
        settings = get_settings()
        api_key = (
            settings.gemini_api_key
            or settings.google_api_key
            or os.getenv("GEMINI_API_KEY")
            or os.getenv("GOOGLE_API_KEY")
        )
        if not api_key:
            return ""

        from langchain_google_genai import ChatGoogleGenerativeAI
        from langchain_core.messages import HumanMessage
        import base64

        vision_llm = ChatGoogleGenerativeAI(
            model="gemini-3.6-flash",
            google_api_key=api_key,
            temperature=0.1,
        )
        b64_data = base64.b64encode(image_bytes).decode("utf-8")
        media_type = mime_type if mime_type else "image/png"

        msg = HumanMessage(
            content=[
                {
                    "type": "text",
                    "text": (
                        "Hãy trích xuất chính xác toàn bộ chữ tiếng Việt/tiếng Anh, bảng biểu "
                        "và công thức toán học (dưới dạng mã LaTeX chuẩn) có trong bức ảnh này. "
                        "Chỉ trả về nội dung chữ bóc tách được."
                    ),
                },
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{media_type};base64,{b64_data}"},
                },
            ]
        )
        resp = vision_llm.invoke([msg])
        content_val = resp.content
        if isinstance(content_val, list):
            texts = [
                part["text"] if isinstance(part, dict) and "text" in part else str(part)
                for part in content_val
            ]
            return "".join(texts).strip()
        return str(content_val).strip()
    except Exception as e:
        logger.warning(f"OCR Vision extraction failed for image chunk: {e}")
        return ""


def _extract_text_from_bytes(file_bytes: bytes, file_name: str, mime_type: str = "") -> str:
    """Extract plain text from file bytes (PDF, DOCX, PPTX, TXT, MD, etc.)."""
    ext = os.path.splitext(file_name)[1].lower()
    text_content = ""

    if ext == ".pdf" or "pdf" in mime_type:
        try:
            reader = PdfReader(io.BytesIO(file_bytes))
            pages_text = []
            for i, page in enumerate(reader.pages):
                extracted = page.extract_text()
                if extracted:
                    pages_text.append(extracted)
            text_content = "\n\n".join(pages_text)
        except Exception as e:
            logger.error(f"Error parsing PDF file {file_name}: {e}")
            raise ValueError(f"Could not parse PDF: {e}") from e
    elif ext in [".docx", ".doc"] or "word" in mime_type:
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                xml_content = z.read("word/document.xml")
                tree = ET.fromstring(xml_content)
                texts = [node.text for node in tree.iter() if node.tag.endswith("}t") and node.text]

                # Extract image OCR for images embedded in Word document
                media_files = [
                    f for f in z.namelist()
                    if f.startswith("word/media/") and not f.endswith("/")
                ]
                for mf in media_files:
                    try:
                        img_bytes = z.read(mf)
                        ocr_text = _ocr_extract_text_from_image(img_bytes)
                        if ocr_text:
                            texts.append(f"\n[Nội dung từ hình ảnh trong tài liệu Word]: {ocr_text}")
                    except Exception:
                        pass

                text_content = " ".join(texts)
        except Exception as e:
            logger.warning(f"Could not parse docx file {file_name}: {e}")
            try:
                text_content = file_bytes.decode("utf-8", errors="ignore")
            except Exception:
                text_content = ""
    elif ext in [".pptx", ".ppt"] or "presentation" in mime_type or "powerpoint" in mime_type:
        try:
            import re
            slide_texts = []
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                slide_names = [
                    n for n in z.namelist()
                    if n.startswith("ppt/slides/slide") and n.endswith(".xml")
                ]
                def _slide_num(n: str) -> int:
                    m = re.search(r"slide(\d+)\.xml$", n)
                    return int(m.group(1)) if m else 999999
                slide_names.sort(key=_slide_num)

                for name in slide_names:
                    try:
                        xml_content = z.read(name)
                        tree = ET.fromstring(xml_content)
                        paragraphs = []
                        for p_node in tree.iter():
                            if p_node.tag.endswith("}p"):
                                p_text = "".join(node.text for node in p_node.iter() if node.tag.endswith("}t") and node.text)
                                if p_text.strip():
                                    paragraphs.append(p_text.strip())

                        # Extract image OCR for images attached to this slide
                        slide_num_str = re.search(r"slide(\d+)\.xml$", name)
                        if slide_num_str:
                            rels_name = f"ppt/slides/_rels/slide{slide_num_str.group(1)}.xml.rels"
                            if rels_name in z.namelist():
                                try:
                                    rels_xml = z.read(rels_name)
                                    rels_tree = ET.fromstring(rels_xml)
                                    for rel in rels_tree.iter():
                                        target = rel.attrib.get("Target", "")
                                        if "media/" in target:
                                            media_path = f"ppt/media/{os.path.basename(target)}"
                                            if media_path in z.namelist():
                                                img_bytes = z.read(media_path)
                                                ocr_text = _ocr_extract_text_from_image(img_bytes)
                                                if ocr_text:
                                                    paragraphs.append(f"[Nội dung từ hình ảnh trên slide]: {ocr_text}")
                                except Exception as rel_err:
                                    logger.debug(f"Could not parse rels for slide {name}: {rel_err}")

                        if paragraphs:
                            slide_texts.append("\n".join(paragraphs))
                    except Exception:
                        pass
            text_content = "\n\n".join(slide_texts)
        except Exception as e:
            logger.warning(f"Could not parse pptx file {file_name}: {e}")
            try:
                text_content = file_bytes.decode("utf-8", errors="ignore")
            except Exception:
                text_content = ""
    elif ext in [".jpg", ".jpeg", ".png", ".webp"] or "image" in mime_type:
        try:
            mtype = mime_type if mime_type else f"image/{ext.replace('.', '')}"
            ocr_res = _ocr_extract_text_from_image(file_bytes, mime_type=mtype)
            text_content = f"[Nội dung trích xuất từ hình ảnh {file_name}]:\n{ocr_res}" if ocr_res else f"Tệp hình ảnh: {file_name}"
        except Exception as e:
            logger.warning(f"Could not parse image file {file_name}: {e}")
            text_content = ""
    elif ext == ".csv" or "csv" in mime_type:
        try:
            import csv
            decoded = file_bytes.decode("utf-8-sig", errors="ignore")
            lines = decoded.splitlines()
            if lines:
                reader = csv.reader(lines)
                rows = list(reader)
                if rows:
                    header = [str(c) for c in rows[0]]
                    md_rows = [f"| {' | '.join(header)} |", f"| {' | '.join(['---']*len(header))} |"]
                    for row in rows[1:500]:
                        r_str = [str(c) for c in row]
                        md_rows.append(f"| {' | '.join(r_str)} |")
                    text_content = f"### Bảng dữ liệu CSV: {file_name}\n" + "\n".join(md_rows)
                else:
                    text_content = decoded
            else:
                text_content = decoded
        except Exception as e:
            logger.warning(f"Could not parse CSV file {file_name}: {e}")
            text_content = file_bytes.decode("utf-8", errors="ignore")
    elif ext in [".xlsx", ".xls"] or "excel" in mime_type or "spreadsheet" in mime_type:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet_texts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                rows = [r for r in rows if any(cell is not None for cell in r)]
                if not rows:
                    continue
                header = [str(c) if c is not None else "" for c in rows[0]]
                md_rows = [f"| {' | '.join(header)} |", f"| {' | '.join(['---']*len(header))} |"]
                for row in rows[1:300]:
                    r_str = [str(c) if c is not None else "" for c in row]
                    md_rows.append(f"| {' | '.join(r_str)} |")
                sheet_texts.append(f"### Sheet: {sheet_name}\n" + "\n".join(md_rows))
            text_content = f"### Tài liệu Excel: {file_name}\n\n" + "\n\n".join(sheet_texts)
        except Exception as e:
            logger.warning(f"Could not parse Excel file {file_name}: {e}")
            text_content = file_bytes.decode("utf-8", errors="ignore")
    else:
        # Fallback to plain text decoding
        try:
            text_content = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text_content = file_bytes.decode("latin-1")
            except Exception as e:
                logger.error(f"Error decoding text file {file_name}: {e}")
                raise ValueError(f"Could not decode text file: {e}") from e

    return text_content.strip()


class RAGService:
    @staticmethod
    async def ingest_document_background(
        course_id: str,
        material_id: str,
        file_bytes: bytes,
        file_name: str,
        object_key: str,
        mime_type: str = "",
    ) -> None:
        """Background task to extract text, chunk, embed, and store into ChromaDB."""
        logger.info(f"Starting background RAG ingestion for material {material_id} ({file_name})...")
        try:
            # 1. Update status to 'processing'
            async with AsyncSessionLocal() as db:
                await MaterialRepository.update_material_status(db, material_id, "processing")

            # 2. Extract text
            raw_text = _extract_text_from_bytes(file_bytes, file_name, mime_type)
            if not raw_text:
                logger.warning(f"No text extracted from material {material_id} ({file_name}).")
                async with AsyncSessionLocal() as db:
                    await MaterialRepository.update_material_status(db, material_id, "completed")
                return

            # 3. Chunk text
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,
                chunk_overlap=200,
                length_function=len,
            )
            chunks = text_splitter.split_text(raw_text)

            # 4. Prepare Documents with Metadata
            documents = []
            for idx, chunk in enumerate(chunks):
                doc = Document(
                    page_content=chunk,
                    metadata={
                        "course_id": course_id,
                        "material_id": material_id,
                        "file_name": file_name,
                        "object_key": object_key,
                        "chunk_index": idx,
                    },
                )
                documents.append(doc)

            # 5. Persist into Vector Store (Qdrant or ChromaDB)
            vector_store = _get_vector_store()
            vector_store.add_documents(documents)
            logger.info(
                f"Successfully ingested {len(documents)} vectors for material {material_id}."
            )

            # 6. Update status to 'completed'
            async with AsyncSessionLocal() as db:
                await MaterialRepository.update_material_status(db, material_id, "completed")

        except Exception as e:
            logger.error(f"Failed RAG ingestion for material {material_id}: {e}", exc_info=True)
            async with AsyncSessionLocal() as db:
                await MaterialRepository.update_material_status(db, material_id, "failed")

    @staticmethod
    def delete_material_vectors(material_id: str) -> None:
        """Delete all vector chunks matching material_id from Qdrant Cloud."""
        try:
            settings = get_settings()
            if QdrantClient is not None and models is not None:
                client = _get_qdrant_client()
                collection_name = settings.qdrant_collection_name or "course_materials"
                client.delete(
                    collection_name=collection_name,
                    points_selector=models.FilterSelector(
                        filter=models.Filter(
                            must=[
                                models.FieldCondition(
                                    key="metadata.material_id",
                                    match=models.MatchValue(value=material_id),
                                )
                            ]
                        )
                    ),
                )
                logger.info(f"Deleted vector chunks for material_id {material_id} from Qdrant.")
        except Exception as e:
            logger.warning(f"Could not delete vectors for material {material_id}: {e}")

    @staticmethod
    def delete_assignment_vectors(assignment_id: str) -> None:
        """Delete all vector chunks matching assignment_id from Qdrant."""
        try:
            settings = get_settings()
            if QdrantClient is not None and models is not None:
                client = _get_qdrant_client()
                collection_name = settings.qdrant_collection_name or "course_materials"
                client.delete(
                    collection_name=collection_name,
                    points_selector=models.FilterSelector(
                        filter=models.Filter(
                            must=[
                                models.FieldCondition(
                                    key="metadata.assignment_id",
                                    match=models.MatchValue(value=assignment_id),
                                )
                            ]
                        )
                    ),
                )
                logger.info(f"Deleted vector chunks for assignment_id {assignment_id} from Qdrant.")
        except Exception as e:
            logger.warning(f"Could not delete vectors for assignment {assignment_id}: {e}")

    @staticmethod
    async def ingest_assignment_background(
        course_id: str,
        assignment_id: str,
        title: str,
        description: str | None = None,
        questions: list[dict[str, Any]] | None = None,
        checklists: list[dict[str, Any]] | None = None,
        file_bytes: bytes | None = None,
        file_name: str | None = None,
        object_key: str | None = None,
        mime_type: str = "",
    ) -> None:
        """Background task to extract text from assignment details, questions, checklists, and attached file, chunk, embed, and store into Qdrant."""
        logger.info(f"Starting background RAG ingestion for assignment {assignment_id} ({title})...")
        try:
            # 1. Clean up existing vectors for this assignment first to avoid duplication
            RAGService.delete_assignment_vectors(assignment_id)

            # 2. Build structured text representation of the assignment
            sections = [f"### YÊU CẦU & NỘI DUNG BÀI TẬP: {title}"]
            sections.append(f"Mã môn học: {course_id} | Mã bài tập: {assignment_id}")

            if description and description.strip():
                sections.append(f"#### Mô tả & Yêu cầu bài tập:\n{description.strip()}")

            if checklists:
                chk_texts = []
                for c in checklists:
                    c_title = c.get("title") or (getattr(c, "title", "") if hasattr(c, "title") else "")
                    c_desc = c.get("description") or (getattr(c, "description", "") if hasattr(c, "description") else "")
                    if c_title:
                        chk_texts.append(f"- {c_title}: {c_desc}".strip(": "))
                if chk_texts:
                    sections.append("#### Danh sách hạng mục cần hoàn thành (Checklists):\n" + "\n".join(chk_texts))

            if questions:
                q_texts = []
                for idx, q in enumerate(questions, start=1):
                    q_text = q.get("question_text") or (getattr(q, "question_text", "") if hasattr(q, "question_text") else "")
                    q_pts = q.get("points", 0) or (getattr(q, "points", 0) if hasattr(q, "points") else 0)
                    q_exp = q.get("expected_answer") or (getattr(q, "expected_answer", None) if hasattr(q, "expected_answer") else None)
                    opts = q.get("options") or (getattr(q, "options", []) if hasattr(q, "options") else [])

                    item_str = f"Câu {idx} ({q_pts} điểm): {q_text}"
                    if opts:
                        opt_strs = []
                        for opt in opts:
                            o_txt = opt.get("option_text") if isinstance(opt, dict) else getattr(opt, "option_text", str(opt))
                            opt_strs.append(f"  + {o_txt}")
                        item_str += "\n" + "\n".join(opt_strs)
                    if q_exp:
                        item_str += f"\n  (Gợi ý/Đáp án tham khảo: {q_exp})"
                    q_texts.append(item_str)

                if q_texts:
                    sections.append("#### Bộ câu hỏi bài tập:\n" + "\n\n".join(q_texts))

            # 3. Extract text from attached file if present
            if file_bytes and file_name:
                try:
                    attachment_text = _extract_text_from_bytes(file_bytes, file_name, mime_type)
                    if attachment_text:
                        sections.append(f"#### Nội dung chi tiết từ tệp đính kèm bài tập ({file_name}):\n{attachment_text}")
                except Exception as file_err:
                    logger.warning(f"Could not extract text from assignment attachment {file_name}: {file_err}")

            full_text = "\n\n".join(sections).strip()
            if not full_text:
                logger.warning(f"No text extracted for assignment {assignment_id}.")
                return

            # 4. Chunk text
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,
                chunk_overlap=200,
                length_function=len,
            )
            chunks = text_splitter.split_text(full_text)

            # 5. Prepare Documents with Metadata
            documents = []
            for idx, chunk in enumerate(chunks):
                doc = Document(
                    page_content=chunk,
                    metadata={
                        "course_id": course_id,
                        "assignment_id": assignment_id,
                        "material_id": assignment_id,
                        "type": "assignment",
                        "file_name": file_name or f"Assignment: {title}",
                        "title": title,
                        "object_key": object_key or "",
                        "chunk_index": idx,
                    },
                )
                documents.append(doc)

            # 6. Persist into Vector Store (Qdrant)
            vector_store = _get_vector_store()
            vector_store.add_documents(documents)
            logger.info(f"Successfully ingested {len(documents)} vectors for assignment {assignment_id} ({title}).")

        except Exception as e:
            logger.error(f"Failed RAG ingestion for assignment {assignment_id}: {e}", exc_info=True)

    @staticmethod
    def search_course_materials(
        course_id: str | None = None,
        query: str = "",
        material_id: str | None = None,
        assignment_id: str | None = None,
        top_k: int | None = None,
    ) -> list[dict[str, Any]]:
        """Perform similarity search or direct fallback retrieval on Qdrant vector store."""
        if not query.strip():
            return []

        settings = get_settings()
        k = top_k if top_k is not None else settings.rag_top_k

        if assignment_id:
            search_filter = models.Filter(
                must=[models.FieldCondition(key="metadata.assignment_id", match=models.MatchValue(value=assignment_id))]
            )
        elif material_id:
            search_filter = models.Filter(
                must=[models.FieldCondition(key="metadata.material_id", match=models.MatchValue(value=material_id))]
            )
        elif course_id:
            search_filter = models.Filter(
                must=[models.FieldCondition(key="metadata.course_id", match=models.MatchValue(value=course_id))]
            )
        else:
            search_filter = None

        formatted_results = []
        try:
            vector_store = _get_vector_store()
            results = vector_store.similarity_search_with_score(
                query=query,
                k=k,
                filter=search_filter,
            )
            for doc, score in results:
                formatted_results.append({
                    "content": doc.page_content,
                    "metadata": doc.metadata or {},
                    "score": float(score),
                })
        except Exception as e:
            logger.error(f"Error searching course materials (course_id={course_id}, material_id={material_id}): {e}")

        summary_keywords = [
            "tóm tắt", "tom tat", "summary", "summarize", "nội dung", "noi dung",
            "overview", "bài giảng", "bai giang", "trích xuất", "trich xuat",
            "câu hỏi", "cau hoi", "ôn tập", "on tap", "bài tập", "bai tap",
            "đề thi", "de thi", "khái niệm", "khai niem", "giải thích", "giai thich",
            "điểm chính", "diem chinh", "ưu tiên", "uu tien"
        ]
        is_summary_query = any(kw in query.lower() for kw in summary_keywords)

        # Fallback: If similarity search returned 0 results or query is a summary request, do direct vector fetch via Qdrant scroll
        if search_filter is not None and (is_summary_query or not formatted_results):
            try:
                if QdrantClient is not None:
                    client = _get_qdrant_client()
                    collection_name = settings.qdrant_collection_name or "course_materials"
                    scroll_res, _ = client.scroll(
                        collection_name=collection_name,
                        scroll_filter=search_filter,
                        limit=k,
                        with_payload=True,
                        with_vectors=False,
                    )
                    existing_contents = {r["content"] for r in formatted_results}
                    for point in scroll_res:
                        payload = point.payload or {}
                        page_content = payload.get("page_content", "") or payload.get("text", "")
                        meta_dict = payload.get("metadata", {})
                        if page_content and page_content not in existing_contents:
                            formatted_results.append({
                                "content": page_content,
                                "metadata": meta_dict or {},
                                "score": 1.0,
                            })
            except Exception as get_err:
                logger.warning(f"Direct fallback retrieval failed for filter {search_filter}: {get_err}")

        return formatted_results

